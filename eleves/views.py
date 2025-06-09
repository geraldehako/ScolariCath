from pyexpat.errors import messages
from django.shortcuts import render, get_object_or_404, redirect
import pandas as pd
from authentifications.decorators import fonctionnalite_autorisee
from cores.models import AnneeScolaires
from eleves.forms import ImportElevesForm, LienParenteForm, LienParenteFormSet
from etablissements.models import Classes, EmploiTemps, Etablissements, Niveaux
from notes.models import Bulletins, Notes
from scolarites.models import Echeances, ModalitePaiements, Paiements, PaiementsCantines, PaiementsTransports
from .models import Eleves, Inscriptions, LienParente, Mutations, Parents, Relances, Scolarites
from .forms import EleveForm, FormulaireInscription, FormulaireInscriptionprimaire, ParentForm
import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from eleves.models import Eleves
from django.core.files.storage import default_storage
from django.utils.dateparse import parse_date
from authentifications.models import Utilisateurs, Roles  # Si nécessaire
from django.core.mail import send_mail 
from django.conf import settings
from django.contrib.auth.hashers import make_password
import logging
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.http import FileResponse
from django.db.models import Count

# === LES ELEVES =========================================================================================================================================
# Back office --------------------------------------------------------------------------------------------------------
@fonctionnalite_autorisee('liste_eleves_back') 
def liste_eleves_back(request):
    eleves = Eleves.objects.all()
    return render(request, 'backoffice/eleves/liste.html', {'eleves': eleves}) 

def liste_eleves_systeme(request):
    eleves = Eleves.objects.all()
    return render(request, 'backoffice/eleves/liste_systeme.html', {'eleves': eleves})

def import_eleves_back(request):
    if request.method == 'POST' and request.FILES.get('fichier'):
        fichier = request.FILES['fichier']
        chemin = default_storage.save('tmp/' + fichier.name, fichier)
        wb = openpyxl.load_workbook(default_storage.path(chemin))
        ws = wb.active

        lignes_importees = 0
        erreurs = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                matricule, nom, prenoms, sexe, naissance, lieu, id_parent = row

                matricule = matricule if matricule else None  # Laisser vide pour génération

                if matricule and Eleves.objects.filter(matricule=matricule).exists():
                    raise ValueError(f"Élève avec matricule '{matricule}' déjà existant.")

                parent = Parents.objects.filter(pk=id_parent).first() if id_parent else None
                date_naissance = parse_date(str(naissance)) if naissance else None
                if not date_naissance:
                    raise ValueError("La date de naissance est requise.")

                Eleves.objects.create(
                    matricule=matricule,
                    nom=nom,
                    prenoms=prenoms,
                    sexe=sexe,
                    date_naissance=date_naissance,
                    lieu_naissance=lieu,
                    parent=parent
                )
                lignes_importees += 1

            except Exception as e:
                erreurs.append(f"Ligne {i+2} : {e}")

        if erreurs:
            messages.warning(request, f"{len(erreurs)} erreur(s) détectée(s) lors de l'importation.")
            for erreur in erreurs:
                messages.error(request, erreur)
        else:
            messages.success(request, f"{lignes_importees} élèves importés avec succès.")

        return redirect('liste_eleves')  # redirection après import

    return render(request, 'backoffice/eleves/import_eleves_form.html')

def ajouter_eleve_back(request):
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES)
        formset = LienParenteFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            eleve = form.save()
            formset.instance = eleve
            formset.save()
            return redirect('liste_eleves')
    else:
        form = EleveForm()
        formset = LienParenteFormSet()
    return render(request, 'backoffice/eleves/formulaire.html', {'form': form, 'formset': formset, 'titre': "Ajouter un élève"})


# front office --------------------------------------------------------------------------------------------------
def liste_eleves(request):
    eleves = Eleves.objects.all()
    return render(request, 'backoffice/eleves/liste.html', {'eleves': eleves})

# apres la premiere annee
def liste_eleves_etablissement(request):
    etablissement = request.user.etablissement

    # Sélectionner les élèves qui ont une inscription dans une classe de cet établissement
    eleves = Eleves.objects.filter(
        inscriptions__classe__etablissement=etablissement
    ).distinct()#.values('matricule', 'nom', 'prenoms','sexe')

    return render(request, 'frontoffice/eleves/liste_eleve_base.html', {
        'eleves': eleves
    })
    
# debut application    
@fonctionnalite_autorisee('liste_eleves_etablissement_start') 
def liste_eleves_etablissement_start(request):
    etablissement = request.user.etablissement

    # Sélectionner les élèves qui ont une inscription dans une classe de cet établissement
    eleves = Eleves.objects.filter(
        origine=etablissement
    ).distinct()#.values('matricule', 'nom', 'prenoms','sexe')

    return render(request, 'frontoffice/eleves/liste_eleve_base.html', {
        'eleves': eleves
    })
    
    
#  Avec les parents ---------------------------------------------------------------------
logger = logging.getLogger(__name__)

def ajouter_eleve(request):
    if request.method == 'POST':
        eleve_form = EleveForm(request.POST, request.FILES)
        parent_form = ParentForm(request.POST)
        formset = LienParenteFormSet(request.POST)

        telephone = request.POST.get('telephone')
        email = request.POST.get('email')

        parent_existant = Parents.objects.filter(Q(telephone=telephone) | Q(email=email)).first()

        if parent_existant:
            parent = parent_existant
            parent_form_valid = True
            nom_complet = parent.nom_complet
            nom_parts = nom_complet.split()
            prenom = nom_parts[0]
            nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
        else:
            parent_form_valid = parent_form.is_valid()
            if parent_form_valid:
                parent = parent_form.save(commit=False)
                telephone = parent_form.cleaned_data.get('telephone')
                email = parent_form.cleaned_data.get('email')
                nom_complet = parent_form.cleaned_data.get('nom_complet')
                nom_parts = nom_complet.split()
                prenom = nom_parts[0]
                nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
            else:
                parent = None

        if eleve_form.is_valid() and parent_form_valid and formset.is_valid():
            try:
                if not parent_existant:
                    parent.save()

                # Création du compte utilisateur pour le parent
                role=Roles.objects.get_or_create(nom="Parents")[0]
                user, created = Utilisateurs.objects.get_or_create(username=telephone, defaults={
                    'first_name': prenom,
                    'last_name': nom,
                    'email': email or '',
                    'password': make_password(telephone),
                    'telephone': telephone,
                    'role': role 
                })

                if created:
                    user.save()

                # ⚠️ Rattacher le compte utilisateur au parent
                parent.utilisateur = user
                parent.save()

                # Création de l'élève
                eleve = eleve_form.save(commit=False)
                eleve.parent = parent
                eleve.save()

                formset.instance = eleve
                formset.save()

                # (Optionnel) Envoi email ou SMS ici

                return redirect('liste_eleves')
            except Exception as e:
                logger.error(f"Erreur lors de la création de l'élève ou du parent : {str(e)}")
        else:
            logger.warning("Formulaires invalides :")
            logger.warning(eleve_form.errors)
            logger.warning(parent_form.errors)
            logger.warning(formset.errors)

    else:
        eleve_form = EleveForm()
        parent_form = ParentForm()
        formset = LienParenteFormSet()

    return render(request, 'frontoffice/eleves/formulaire.html', {
        'form': eleve_form,
        'form_parent': parent_form,
        'formset': formset,
        'titre': "Ajouter un élève"
    })
    
def ajouter_eleve_etablissement(request):
    if request.method == 'POST':
        eleve_form = EleveForm(request.POST, request.FILES)
        parent_form = ParentForm(request.POST)
        formset = LienParenteFormSet(request.POST)

        telephone = request.POST.get('telephone')
        email = request.POST.get('email')

        parent_existant = Parents.objects.filter(Q(telephone=telephone) | Q(email=email)).first()

        if parent_existant:
            parent = parent_existant
            parent_form_valid = True
            nom_complet = parent.nom_complet
            nom_parts = nom_complet.split()
            prenom = nom_parts[0]
            nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
        else:
            parent_form_valid = parent_form.is_valid()
            if parent_form_valid:
                parent = parent_form.save(commit=False)
                telephone = parent_form.cleaned_data.get('telephone')
                email = parent_form.cleaned_data.get('email')
                nom_complet = parent_form.cleaned_data.get('nom_complet')
                nom_parts = nom_complet.split()
                prenom = nom_parts[0]
                nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
            else:
                parent = None

        if eleve_form.is_valid() and parent_form_valid and formset.is_valid():
            try:
                if not parent_existant:
                    parent.save()

                # Création du compte utilisateur pour le parent
                role=Roles.objects.get_or_create(nom="Parents")[0]
                user, created = Utilisateurs.objects.get_or_create(username=telephone, defaults={
                    'first_name': prenom,
                    'last_name': nom,
                    'email': email or '',
                    'password': make_password(telephone),
                    'telephone': telephone,
                    'role': role 
                })

                if created:
                    user.save()

                # ⚠️ Rattacher le compte utilisateur au parent
                parent.utilisateur = user
                parent.save()

                # Création de l'élève
                eleve = eleve_form.save(commit=False)
                eleve.parent = parent
                eleve.save()

                formset.instance = eleve
                formset.save()

                # (Optionnel) Envoi email ou SMS ici

                return redirect('liste_eleves')
            except Exception as e:
                logger.error(f"Erreur lors de la création de l'élève ou du parent : {str(e)}")
        else:
            logger.warning("Formulaires invalides :")
            logger.warning(eleve_form.errors)
            logger.warning(parent_form.errors)
            logger.warning(formset.errors)

    else:
        eleve_form = EleveForm()
        parent_form = ParentForm()
        formset = LienParenteFormSet()

    return render(request, 'frontoffice/eleves/formulaire.html', {
        'form': eleve_form,
        'form_parent': parent_form,
        'formset': formset,
        'titre': "Ajouter un élève"
    })

from django.db import transaction


@transaction.atomic
def modifier_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleves, pk=eleve_id)
    parent = eleve.parent

    parent_partage = Eleves.objects.filter(parent=parent).exclude(pk=eleve.pk).exists()

    if request.method == 'POST':
        eleve_form = EleveForm(request.POST, request.FILES, instance=eleve)
        formset = LienParenteFormSet(request.POST, instance=eleve)

        # Si le parent est partagé, ne pas recharger le form avec POST (empêche modification)
        parent_form = ParentForm(instance=parent) if parent_partage else ParentForm(request.POST, instance=parent)

        if eleve_form.is_valid() and formset.is_valid() and (parent_partage or parent_form.is_valid()):
            eleve = eleve_form.save()

            if not parent_partage:
                parent = parent_form.save()

                # Mise à jour de l'utilisateur lié
                if parent.utilisateur:
                    nom_parts = parent.nom_complet.split()
                    prenom = nom_parts[0]
                    nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
                    user = parent.utilisateur
                    user.first_name = prenom
                    user.last_name = nom
                    user.username = parent.telephone
                    user.email = parent.email
                    user.save()

            formset.save()

            messages.success(request, "L'élève a été modifié avec succès.")
            return redirect('liste_eleves')
        else:
            messages.error(request, "Veuillez corriger les erreurs dans le formulaire.")
    else:
        eleve_form = EleveForm(instance=eleve)
        parent_form = ParentForm(instance=parent)
        formset = LienParenteFormSet(instance=eleve)

    return render(request, 'frontoffice/eleves/formulaire.html', {
        'form': eleve_form,
        'form_parent': parent_form,
        'formset': formset,
        'titre': f"Modifier l'élève {eleve.nom} {eleve.prenoms}",
        'parent_partage': parent_partage,  # pour désactiver les champs côté template
    })




#  Sans les parents ---------------------------------------------------------------------
from django.utils.dateparse import parse_date

def ajouter_elevee(request):
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES)
        if form.is_valid():
            eleve = form.save(commit=False)

            # Nettoyage éventuel de la date de naissance (si tu veux parser manuellement)
            naissance = form.cleaned_data.get('date_naissance')
            eleve.date_naissance = parse_date(str(naissance)) if naissance else None

            eleve.save()  # Le matricule est généré automatiquement si vide

            return redirect('eleves_liste')
    else:
        form = EleveForm()
    return render(request, 'frontoffice/eleves/formulaire.html', {'form': form, 'titre': "Ajouter un élève"})

# DOSSIER ELEVES --------------------------------------------------------------------------------------------------------------------
@fonctionnalite_autorisee('detail_eleve')
def detail_eleve(request, matricule):
    eleve = get_object_or_404(Eleves, matricule=matricule)
    inscriptions = Inscriptions.objects.filter(eleve=eleve).select_related('classe', 'annee_scolaire')
    lienparent = LienParente.objects.filter(eleve__matricule=matricule).select_related('parent__utilisateur')
    paiements = Paiements.objects.filter(inscription__eleve=eleve).select_related('inscription')
    paiementtransports = PaiementsTransports.objects.filter(inscription__eleve=eleve).select_related('inscription')
    paiementcantines = PaiementsCantines.objects.filter(inscription__eleve=eleve).select_related('inscription')
    notes = Notes.objects.filter(eleve=eleve).select_related('matiere', 'periode')
    inscription_active = inscriptions.first()  # ou filtre selon l'année active
    print(lienparent)
    context = {
        'eleve': eleve,
        'inscriptions': inscriptions,
        'parents': lienparent,
        'paiements': paiements,
        'paiements_transport' : paiementtransports,
        'paiements_cantine' : paiementcantines,
        'notes': notes,
        'inscription_active': inscription_active,
    }
    return render(request, 'frontoffice/eleves/detail_eleve.html', context)


def modifier_elevee(request, eleve_id):
    eleve = get_object_or_404(Eleves, id=eleve_id)
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        if form.is_valid():
            form.save()
            return redirect('liste_eleves')
    else:
        form = EleveForm(instance=eleve)
    return render(request, 'frontoffice/eleves/formulaire.html', {'form': form, 'titre': "Modifier l'élève"})

def supprimer_eleve(request, eleve_id):
    eleve = get_object_or_404(Eleves, id=eleve_id)
    if request.method == 'POST':
        eleve.delete()
        return redirect('liste_eleves')
    return render(request, 'frontoffice/eleves/confirmation_suppression.html', {'eleve': eleve})


def import_eleves(request): 
    if request.method == 'POST' and request.FILES.get('fichier'):
        fichier = request.FILES['fichier']
        chemin = default_storage.save('tmp/' + fichier.name, fichier)
        wb = openpyxl.load_workbook(default_storage.path(chemin))
        ws = wb.active

        lignes_importees = 0
        erreurs = []

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
            try:
                matricule, nom, prenoms, sexe, naissance, lieu, id_parent = row

                matricule = matricule if matricule else None  # Laisser vide pour génération

                if matricule and Eleves.objects.filter(matricule=matricule).exists():
                    raise ValueError(f"Élève avec matricule '{matricule}' déjà existant.")

                parent = Parents.objects.filter(pk=id_parent).first() if id_parent else None
                date_naissance = parse_date(str(naissance)) if naissance else None
                if not date_naissance:
                    raise ValueError("La date de naissance est requise.")

                Eleves.objects.create(
                    matricule=matricule,
                    nom=nom,
                    prenoms=prenoms,
                    sexe=sexe,
                    date_naissance=date_naissance,
                    lieu_naissance=lieu,
                    parent=parent
                )
                lignes_importees += 1

            except Exception as e:
                erreurs.append(f"Ligne {i+2} : {e}")

        if erreurs:
            messages.warning(request, f"{len(erreurs)} erreur(s) détectée(s) lors de l'importation.")
            for erreur in erreurs:
                messages.error(request, erreur)
        else:
            messages.success(request, f"{lignes_importees} élèves importés avec succès.")

        return redirect('liste_eleves')  # redirection après import

    return render(request, 'backoffice/eleves/import_eleves_form.html')



# === LES PARENTS ====================================================================================================================================
@fonctionnalite_autorisee('liste_liens_back')
def liste_liens_back(request):
    liens = LienParente.objects.select_related('eleve', 'parent').all()
    return render(request, 'backoffice/parents/liste.html', {'liens': liens})


def liste_liens(request):
    liens = LienParente.objects.select_related('eleve', 'parent').all()
    return render(request, 'frontoffice/liens/liste.html', {'liens': liens})

def ajouter_lien(request):
    if request.method == 'POST':
        form = LienParenteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('liste_liens')
    else:
        form = LienParenteForm()
    return render(request, 'frontoffice/liens/formulaire.html', {'form': form, 'titre': 'Ajouter un lien de parenté'})

def modifier_lien(request, pk):
    lien = get_object_or_404(LienParente, pk=pk)
    if request.method == 'POST':
        form = LienParenteForm(request.POST, instance=lien)
        if form.is_valid():
            form.save()
            return redirect('liste_liens')
    else:
        form = LienParenteForm(instance=lien)
    return render(request, 'frontoffice/liens/formulaire.html', {'form': form, 'titre': 'Modifier le lien de parenté'})

def supprimer_lien(request, pk):
    lien = get_object_or_404(LienParente, pk=pk)
    if request.method == 'POST':
        lien.delete()
        return redirect('liste_liens')
    return render(request, 'frontoffice/liens/confirmation_suppression.html', {'lien': lien})


# ===  INSCRIPTION  ==================================================================================================================================================  
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Q
from django.http import JsonResponse

def traitement_matriculeONE(request):
    if request.method == 'POST':
        matricule = request.POST.get('matricule')
        try:
            eleve = Eleves.objects.get(matricule=matricule)
        except Eleves.DoesNotExist:
            return JsonResponse({'action': 'ajouter_eleveprerempli', 'matricule': matricule})

        annee_active = AnneeScolaires.objects.get(active=True)
        inscription = Inscriptions.objects.filter(eleve=eleve, annee_scolaire=annee_active).first()

        if inscription:
            return JsonResponse({'action': 'paiement', 'eleve_id': eleve.id})
        else:
            return JsonResponse({'action': 'inscription', 'eleve_id': eleve.id})
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@login_required
def traitement_matricule(request):
    if request.method == 'POST':
        matricule = request.POST.get('matricule')
        if not matricule:
            return JsonResponse({'error': 'Matricule manquant'}, status=400)

        try:
            eleve = Eleves.objects.get(matricule=matricule)
        except Eleves.DoesNotExist:
            return JsonResponse({'action': 'ajouter_eleveprerempli', 'matricule': matricule})

        annee_active = AnneeScolaires.objects.get(active=True)
        inscription = Inscriptions.objects.filter(eleve=eleve, annee_scolaire=annee_active).first()
        etablissement_utilisateur = request.user.etablissement

        if inscription:
            if inscription.classe.etablissement == etablissement_utilisateur:
                return JsonResponse({'action': 'paiement', 'eleve_id': eleve.id})
            else:
                return JsonResponse({
                    'action': 'deja_inscrit_ailleurs',
                    'message': f"L'élève est déjà inscrit dans l'établissement : {inscription.classe.etablissement.nom}"
                })
        else:
            return JsonResponse({'action': 'inscription', 'eleve_id': eleve.id})

    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@fonctionnalite_autorisee('detail_paiement')
def detail_paiement(request, eleve_id): 
    annee_active = AnneeScolaires.objects.get(active=True)
    eleve = get_object_or_404(Eleves, id=eleve_id)
    inscription = get_object_or_404(Inscriptions, eleve=eleve, annee_scolaire=annee_active)
    paiements = Paiements.objects.filter(inscription=inscription)
    return render(request, 'frontoffice/paiements/detail_paiement.html', {
        'eleve': eleve,
        'inscription': inscription,
        'paiements': paiements,
    })

def ajouter_inscription(request, eleve_id): 
   
    etablissement = request.user.etablissement
    types_etablissement = etablissement.types.all()  # ManyToMany vers Cycles

    # Exemple pour afficher les noms des types
    for cycle in types_etablissement: 
        print(cycle.nom)
    
    eleve = get_object_or_404(Eleves, id=eleve_id)
    annee_active = AnneeScolaires.objects.get(active=True)
    if Inscriptions.objects.filter(eleve=eleve, annee_scolaire=annee_active).exists():
        return redirect('detail_paiement', eleve_id=eleve.id)

    if request.method == 'POST':
        
        if cycle.nom == 'Préscolaire' or cycle.nom == 'Primaire' :
            form = FormulaireInscriptionprimaire(request.POST, user=request.user)
        else :
            form = FormulaireInscription(request.POST, user=request.user)
            
        if form.is_valid():
            inscription = form.save(commit=False)
            inscription.eleve = eleve
            inscription.annee_scolaire = annee_active
            inscription.utilisateur = request.user
            if cycle.nom == 'Préscolaire' or cycle.nom == 'Primaire' :
                inscription.statut = 'non_affecte'
            inscription.save()
            
            # Déterminer la modalité à associer
            if cycle.nom == 'Préscolaire' or cycle.nom == 'Primaire' :
                statut = 'non_affecte'
            else :
                statut = inscription.statut  # 'affecte' ou 'non_affecte'

            if statut == 'non_affecte':
                modalite = ModalitePaiements.objects.filter(
                    etablissement=inscription.classe.etablissement if inscription.classe else request.user.etablissement,
                    annee_scolaire=inscription.annee_scolaire,
                    niveau=inscription.classe.niveau if inscription.classe else None,
                    applicable_aux_non_affectes=True
                ).first()
            else:  # statut == 'affecte'    
                modalite = ModalitePaiements.objects.filter(
                    etablissement=inscription.classe.etablissement if inscription.classe else request.user.etablissement,
                    annee_scolaire=inscription.annee_scolaire,
                    niveau=inscription.classe.niveau if inscription.classe else None
                ).first()

            if modalite:
                Scolarites.objects.create(
                    inscription=inscription,
                    modalite=modalite
                )
                # Étape supplémentaire : création des relances pour chaque échéance
                echeances = Echeances.objects.filter(modalite=modalite)
                for echeance in echeances:
                    montant = echeance.montant
                    relance = Relances.objects.create(
                        inscription=inscription,
                        echeance=echeance,
                        date_relance=echeance.date_limite,
                        statut='active',
                        echeance_montant=montant,
                        total_verse=0,
                        total_solde=montant
                    )
            else:
                messages.warning(request, "Aucune modalité trouvée pour cette inscription.")

            #return redirect('detail_paiement', eleve_id=eleve.id)
            return redirect('ajouter_paiement', inscription_id=inscription.id)

    else:
        if cycle.nom == 'Préscolaire' or cycle.nom == 'Primaire' :
            form = FormulaireInscriptionprimaire(request.POST, user=request.user)
        else :
            form = FormulaireInscription(request.POST, user=request.user)
            
    if cycle.nom == 'Préscolaire' or cycle.nom == 'Primaire' :
        return render(request, 'frontoffice/inscriptions/ajouter_inscription_primaire.html', {'form': form, 'eleve': eleve})
    else :
        return render(request, 'frontoffice/inscriptions/ajouter_inscription.html', {'form': form, 'eleve': eleve})


def ajouter_eleveprerempli(request):
    matricule = request.GET.get('matricule', '')
    if request.method == 'POST':
        eleve_form = EleveForm(request.POST)
        if eleve_form.is_valid():
            eleve = eleve_form.save()
            return redirect('ajouter_inscription', eleve_id=eleve.id)
    else:
        eleve_form = EleveForm(initial={'matricule': matricule})

    return render(request, 'frontoffice/eleves/formulaire.html', {'form': eleve_form})



# Front office -------------------------------------------------------------------------------------
@login_required
def eleves_inscrits_etablissementOK(request):
    user_etablissement = request.user.etablissement  # à adapter selon ton modèle
    annee_active = AnneeScolaires.objects.get(active=True)
    classes_disponibles = Classes.objects.filter(etablissement=request.user.etablissement)


    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        classe__etablissement=user_etablissement
    ).select_related('eleve', 'classe__niveau')

    return render(request, 'frontoffice/eleves/liste_eleves_inscrits.html', {
        'inscriptions': inscriptions,
        'etablissement': user_etablissement,
        'annee': annee_active,
        classes_disponibles:classes_disponibles
    })

@fonctionnalite_autorisee('eleves_inscrits_etablissement')   
def eleves_inscrits_etablissement(request):
    user = request.user
    etablissement = user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    # Toutes les inscriptions pour l’année active et l’établissement
    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active
    ).select_related('eleve', 'classe', 'classe__niveau')

    # Classes disponibles dans cet établissement pour la mutation
    classes_disponibles = Classes.objects.filter(etablissement=etablissement,annee_scolaire=annee_active)

    context = {
        'inscriptions': inscriptions,
        'etablissement': etablissement,
        'annee': annee_active,
        'classes_disponibles': classes_disponibles,
    }
    return render(request, 'frontoffice/eleves/liste_eleves_inscrits.html', context)

@login_required
def export_eleves_inscrits_excel(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active
    ).select_related('eleve', 'classe__niveau')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves inscrits"

    ws.append(['Niveau', 'Classe', 'Nom', 'Prénoms'])

    for ins in inscriptions:
        ws.append([
            ins.classe.niveau.nom,
            ins.classe.nom,
            ins.eleve.nom,
            ins.eleve.prenoms
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=eleves_inscrits.xlsx'
    wb.save(response)
    return response


@login_required
def export_eleves_inscrits_pdf(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active
    ).select_related('eleve', 'classe__niveau')

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, "Liste des élèves inscrits")

    p.setFont("Helvetica", 12)
    y = height - 80
    p.drawString(50, y, "Niveau")
    p.drawString(150, y, "Classe")
    p.drawString(250, y, "Nom")
    p.drawString(400, y, "Prénoms")

    y -= 20
    for ins in inscriptions:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, str(ins.classe.niveau.nom))
        p.drawString(150, y, str(ins.classe.nom))
        p.drawString(250, y, ins.eleve.nom)
        p.drawString(400, y, ins.eleve.prenoms)
        y -= 20

    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='eleves_inscrits.pdf')

@fonctionnalite_autorisee('eleves_inscrits_abandon_etablissement')   
def eleves_inscrits_abandon_etablissement(request):
    user = request.user
    etablissement = user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    # Toutes les inscriptions pour l’année active et l’établissement
    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active, etat="abandon"
    ).select_related('eleve', 'classe', 'classe__niveau')

    # Classes disponibles dans cet établissement pour la mutation , 
    classes_disponibles = Classes.objects.filter(etablissement=etablissement,annee_scolaire=annee_active)

    context = {
        'inscriptions': inscriptions,
        'etablissement': etablissement,
        'annee': annee_active,
        'classes_disponibles': classes_disponibles,
    }
    return render(request, 'frontoffice/eleves/liste_eleves_inscrits_abandon.html', context)

@fonctionnalite_autorisee('eleves_inscrits_abandon_tous')    
def eleves_inscrits_abandon_tous(request): 
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,etat="abandon"
        
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    context = {
        'inscriptions': inscriptions,
        'annee': annee_active,
    }
    return render(request, 'backoffice/eleves/liste_eleves_abandons_tous.html', context)

@fonctionnalite_autorisee('eleves_inscrits_reduction_tous') 
def eleves_inscrits_reduction_tous(request): 
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        reduction__gt=1
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    context = {
        'inscriptions': inscriptions,
        'annee': annee_active,
    }
    return render(request, 'backoffice/eleves/liste_eleves_reduction_tous.html', context)

@login_required
def export_eleves_inscrits_abandon_excel(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active,etat="abandon"
    ).select_related('eleve', 'classe__niveau')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves inscrits"

    ws.append(['Niveau', 'Classe', 'Nom', 'Prénoms'])

    for ins in inscriptions:
        ws.append([
            ins.classe.niveau.nom,
            ins.classe.nom,
            ins.eleve.nom,
            ins.eleve.prenoms
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=eleves_inscrits.xlsx'
    wb.save(response)
    return response


@login_required
def export_eleves_inscrits_abandon_pdf(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active,etat="abandon"
    ).select_related('eleve', 'classe__niveau')

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, "Liste des élèves inscrits")

    p.setFont("Helvetica", 12)
    y = height - 80
    p.drawString(50, y, "Niveau")
    p.drawString(150, y, "Classe")
    p.drawString(250, y, "Nom")
    p.drawString(350, y, "Prénoms")
    p.drawString(400, y, "Total du")
    p.drawString(450, y, "Total payé")
    p.drawString(500, y, "Restant")

    y -= 20
    for ins in inscriptions:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, str(ins.classe.niveau.nom))
        p.drawString(150, y, str(ins.classe.nom))
        p.drawString(250, y, ins.eleve.nom)
        p.drawString(350, y, ins.eleve.prenoms)
        p.drawString(400, y, str(ins.montant_total_du))
        p.drawString(450, y, str(ins.montant_total_paye))
        p.drawString(500, y, str(ins.solde_restant))
        y -= 20

    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='eleves_inscrits.pdf')

def export_eleves_inscrits_abandon_tous_excel(request):
    annee_active = AnneeScolaires.objects.get(active=True)
    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        etat="abandon"
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves abandons"

    # En-têtes
    ws.append(["Établissement", "Niveau", "Classe", "Nom", "Prénoms", "Total dû", "Total payé", "Restant"])

    # Données
    for i in inscriptions:
        ws.append([
            i.classe.etablissement.nom,
            i.classe.niveau.nom,
            i.classe.nom,
            i.eleve.nom,
            i.eleve.prenoms,
            i.montant_total_du(),
            i.montant_total_paye(),
            i.solde_restant()
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=eleves_abandons.xlsx"
    wb.save(response)
    return response

from reportlab.lib.pagesizes import A4, landscape
def export_eleves_inscrits_abandon_tous_pdf(request):
    annee_active = AnneeScolaires.objects.get(active=True)
    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        etat="abandon"
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=eleves_abandons.pdf'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(30, height - 40, f"Liste des élèves abandons - Année {annee_active.libelle}")

    c.setFont("Helvetica-Bold", 10)
    headers = ["Établissement", "Niveau", "Classe", "Nom", "Prénoms", "Total dû", "Total payé", "Restant"]
    y = height - 70
    x = 30

    for header in headers:
        c.drawString(x, y, header)
        x += 90

    y -= 20
    c.setFont("Helvetica", 9)

    for i in inscriptions:
        x = 30
        ligne = [
            i.classe.etablissement.nom,
            i.classe.niveau.nom,
            i.classe.nom,
            i.eleve.nom,
            i.eleve.prenoms,
            str(i.montant_total_du()),
            str(i.montant_total_paye()),
            str(i.solde_restant())
        ]
        for item in ligne:
            c.drawString(x, y, item)
            x += 90
        y -= 18
        if y < 50:
            c.showPage()
            y = height - 70
            c.setFont("Helvetica", 9)

    c.save()
    return response

def export_eleves_inscrits_reduction_tous_excel(request):
    annee_active = AnneeScolaires.objects.get(active=True)
    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        reduction__gt=1
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves abandons"

    # En-têtes
    ws.append(["Établissement", "Niveau", "Classe", "Nom", "Prénoms", "Total dû", "Total payé", "Restant"])

    # Données
    for i in inscriptions:
        ws.append([
            i.classe.etablissement.nom,
            i.classe.niveau.nom,
            i.classe.nom,
            i.eleve.nom,
            i.eleve.prenoms,
            i.montant_total_du(),
            i.montant_total_paye(),
            i.solde_restant()
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=eleves_abandons.xlsx"
    wb.save(response)
    return response

from reportlab.lib.pagesizes import A4, landscape
def export_eleves_inscrits_reduction_tous_pdf(request):
    annee_active = AnneeScolaires.objects.get(active=True)
    inscriptions = Inscriptions.objects.filter(
        annee_scolaire=annee_active,
        reduction__gt=1
    ).select_related('eleve', 'classe', 'classe__niveau', 'classe__etablissement')

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=eleves_abandons.pdf'

    c = canvas.Canvas(response, pagesize=landscape(A4))
    width, height = landscape(A4)

    c.setFont("Helvetica-Bold", 14)
    c.drawString(30, height - 40, f"Liste des élèves abandons - Année {annee_active.libelle}")

    c.setFont("Helvetica-Bold", 10)
    headers = ["Établissement", "Niveau", "Classe", "Nom", "Prénoms", "Total dû", "Total payé", "Restant"]
    y = height - 70
    x = 30

    for header in headers:
        c.drawString(x, y, header)
        x += 90

    y -= 20
    c.setFont("Helvetica", 9)

    for i in inscriptions:
        x = 30
        ligne = [
            i.classe.etablissement.nom,
            i.classe.niveau.nom,
            i.classe.nom,
            i.eleve.nom,
            i.eleve.prenoms,
            str(i.montant_total_du()),
            str(i.montant_total_paye()),
            str(i.solde_restant())
        ]
        for item in ligne:
            c.drawString(x, y, item)
            x += 90
        y -= 18
        if y < 50:
            c.showPage()
            y = height - 70
            c.setFont("Helvetica", 9)

    c.save()
    return response

@fonctionnalite_autorisee('eleves_inscrits_reduction_etablissement')  
def eleves_inscrits_reduction_etablissement(request):
    user = request.user
    etablissement = user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    # Toutes les inscriptions pour l’année active et l’établissement
    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active, reduction__gt=1,
    ).select_related('eleve', 'classe', 'classe__niveau')

    # Classes disponibles dans cet établissement pour la mutation
    classes_disponibles = Classes.objects.filter(etablissement=etablissement,annee_scolaire=annee_active)

    context = {
        'inscriptions': inscriptions,
        'etablissement': etablissement,
        'annee': annee_active,
        'classes_disponibles': classes_disponibles,
    }
    return render(request, 'frontoffice/eleves/liste_eleves_inscrits_reduction.html', context)

@login_required
def export_eleves_inscrits_reduction_excel(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active,reduction__gt=1
    ).select_related('eleve', 'classe__niveau')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Élèves inscrits"

    ws.append(['Niveau', 'Classe', 'Nom', 'Prénoms','Remise'])

    for ins in inscriptions:
        ws.append([
            ins.classe.niveau.nom,
            ins.classe.nom,
            ins.eleve.nom,
            ins.eleve.prenoms,
            ins.reduction
        ])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=eleves_inscrits.xlsx'
    wb.save(response)
    return response


@login_required
def export_eleves_inscrits_reduction_pdf(request):
    user = request.user
    etablissement = request.user.etablissement
    annee_active = AnneeScolaires.objects.get(active=True)

    inscriptions = Inscriptions.objects.filter(
        classe__etablissement=etablissement,
        annee_scolaire=annee_active,reduction__gt=1
    ).select_related('eleve', 'classe__niveau')

    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    p.setFont("Helvetica-Bold", 14)
    p.drawString(50, height - 50, "Liste des élèves inscrits")

    p.setFont("Helvetica", 12)
    y = height - 80
    p.drawString(50, y, "Niveau")
    p.drawString(150, y, "Classe")
    p.drawString(250, y, "Nom")
    p.drawString(400, y, "Prénoms")
    p.drawString(500, y, "Remise")

    y -= 20
    for ins in inscriptions:
        if y < 50:
            p.showPage()
            y = height - 50
        p.drawString(50, y, str(ins.classe.niveau.nom))
        p.drawString(150, y, str(ins.classe.nom))
        p.drawString(250, y, ins.eleve.nom)
        p.drawString(400, y, ins.eleve.prenoms)
        p.drawString(500, y, str(ins.reduction))
        y -= 20

    p.showPage()
    p.save()
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='eleves_inscrits.pdf')

@fonctionnalite_autorisee('classes_avec_eleves_inscrits')
def classes_avec_eleves_inscrits(request): 
    annee_active = AnneeScolaires.objects.get(active=True)
    etablissement = request.user.etablissement

    # Récupérer les classes avec le nombre d’élèves inscrits
    classes = Classes.objects.filter(
        etablissement=etablissement,
        inscriptions__annee_scolaire=annee_active
    ).annotate(
        nb_inscrits=Count('inscriptions', filter=Q(inscriptions__annee_scolaire=annee_active))
    ).filter(nb_inscrits__gt=0)

    return render(request, 'frontoffice/eleves/liste_classes.html', {'classes': classes,'etablissement':etablissement})

from openpyxl import Workbook
@fonctionnalite_autorisee('liste_eleves_inscrits_par_classe')
def liste_eleves_inscrits_par_classe(request, classe_id):
    classe = get_object_or_404(Classes, pk=classe_id, etablissement=request.user.etablissement)
    annee_active = AnneeScolaires.objects.get(active=True)
    
    inscriptions = Inscriptions.objects.filter(
        classe=classe,
        annee_scolaire=annee_active
    ).select_related('eleve')
    
    export_format = request.GET.get("export")

    total = inscriptions.count()
    garcons = inscriptions.filter(eleve__sexe='M').count()
    filles = inscriptions.filter(eleve__sexe='F').count()

    if export_format == "excel":
        return exporter_excel(classe, annee_active, inscriptions, total, garcons, filles)
    elif export_format == "pdf":
        return exporter_pdf(classe, annee_active, inscriptions, total, garcons, filles)

    return render(request, 'frontoffice/eleves/eleves_inscrits_par_classe.html', {
        'classe': classe,
        'inscriptions': inscriptions
    })

def exporter_excel(classe, annee, inscriptions, total, garcons, filles):
    wb = Workbook()
    ws = wb.active
    ws.title = "Élèves"

    # En-têtes
    ws.append([
        f"Établissement : {classe.etablissement.nom}",
        f"Année scolaire : {annee.libelle}",
        f"Classe : {classe.nom}",
        f"Effectif : {total} (G: {garcons}, F: {filles})"
    ])
    ws.append([])

    # Colonnes
    ws.append(['Matricule', 'Nom', 'Prénom', 'Sexe', 'Date de naissance'])

    for ins in inscriptions:
        e = ins.eleve
        ws.append([e.matricule, e.nom, e.prenoms, e.sexe, e.date_naissance.strftime("%d/%m/%Y")])

    # Réponse
    response = HttpResponse(content_type='application/vnd.ms-excel')
    filename = f"eleves_{classe.nom}_{annee.libelle}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

def exporter_pdf(classe, annee, inscriptions, total, garcons, filles):
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    y = height - 50

    c.setFont("Helvetica-Bold", 14)
    c.drawString(50, y, f"Établissement : {classe.etablissement.nom}")
    y -= 20
    c.drawString(50, y, f"Année scolaire : {annee.libelle}")
    y -= 20
    c.drawString(50, y, f"Classe : {classe.nom}")
    y -= 20
    c.drawString(50, y, f"Effectif : {total} (G: {garcons}, F: {filles})")
    y -= 40

    c.setFont("Helvetica-Bold", 12)
    c.drawString(50, y, "Matricule")
    c.drawString(150, y, "Nom")
    c.drawString(250, y, "Prénom")
    c.drawString(350, y, "Sexe")
    c.drawString(420, y, "Naissance")
    y -= 20

    c.setFont("Helvetica", 10)
    for ins in inscriptions:
        if y < 50:
            c.showPage()
            y = height - 50
        e = ins.eleve
        c.drawString(50, y, str(e.matricule))
        c.drawString(150, y, e.nom)
        c.drawString(250, y, e.prenoms)
        c.drawString(350, y, e.sexe)
        c.drawString(420, y, e.date_naissance.strftime("%d/%m/%Y"))
        y -= 18

    c.save()
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f"eleves_{classe.nom}_{annee.libelle}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

 
#  les mutations de classe ---------------------------------------------------------------------------------------------------------------------
from django.shortcuts import redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone

def muter_eleves(request):
    if request.method == 'POST':
        ids_eleves = request.POST.getlist('eleves_selectionnes')
        id_nouvelle_classe = request.POST.get('nouvelle_classe')
        motif = request.POST.get('motif')

        nouvelle_classe = get_object_or_404(Classes, id=id_nouvelle_classe)

        for eleve_id in ids_eleves:
            eleve = get_object_or_404(Eleves, id=eleve_id)
            inscription = Inscriptions.objects.filter(
                eleve=eleve,
                annee_scolaire__active=True
            ).first()
            if inscription and inscription.classe != nouvelle_classe:
                Mutations.objects.create(
                    eleve=eleve,
                    ancienne_classe=inscription.classe,
                    nouvelle_classe=nouvelle_classe,
                    date_mutation=timezone.now().date(),
                    motif=motif
                )
                # Mise à jour de la classe dans l'inscription
                inscription.classe = nouvelle_classe
                inscription.save()

        messages.success(request, "Mutation(s) effectuée(s) avec succès.")
    return redirect('liste_eleves_inscrits')
#  les mutations de classe ---------------------------------------------------------------------------------------------------------------------

#  les relances --------------------------------------------------------------------------------------------------------------------------------
def liste_relances(request, statut):
    relances = Relances.objects.filter(statut=statut).select_related('inscription__eleve', 'echeance')
    return render(request, 'relances/liste_relances.html', {
        'relances': relances,
        'statut': statut
    })

def generer_relances_pour_echeance(echeance):
    inscriptions = Inscriptions.objects.filter(classe__niveau__cycle=echeance.cycle, annee=echeance.annee_scolaire)

    for inscription in inscriptions:
        montant = echeance.montant
        total_verse = sum(
            paiement.montant for paiement in inscription.paiements_set.filter(echeance=echeance, statut_validation='valide')
        )
        solde = max(0, montant - total_verse)

        Relances.objects.update_or_create(
            inscription=inscription,
            echeance=echeance,
            defaults={
                'date_relance': timezone.now().date(),
                'echeance_montant': montant,
                'total_verse': total_verse,
                'total_solde': solde
            }
        )
        

from datetime import date
@fonctionnalite_autorisee('liste_relances_non_a_jour')   
def liste_relances_non_a_jour(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    etablissement = request.user.etablissement
    aujourd_hui = date.today()

    # Filtrer les relances non soldées dont la date de relance est passée ou aujourd'hui
    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        statut='active',
        date_relance__lte=aujourd_hui
    ).select_related('inscription__eleve', 'echeance')

    return render(request, 'frontoffice/relances/liste_relances_non_a_jour.html', {
        'relances': relances,
        'aujourd_hui': aujourd_hui,
    })

from weasyprint import HTML
from django.template.loader import get_template
from django.http import HttpResponse
from django.utils.timezone import now

def export_relances_pdf(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    etablissement = request.user.etablissement
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance')

    template = get_template('frontoffice/relances/export_relances_pdf.html')
    html_string = template.render({
        'relances': relances,
        'date_du_jour': today,
        'etablissement': etablissement,
        'annee': annee_active,
    })

    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relances_non_ajournees.pdf"'
    return response

import openpyxl
from django.http import HttpResponse

def export_relances_excel(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True) 
    etablissement = request.user.etablissement
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relances non à jour"

    # En-têtes
    ws.append(["Élève", "Statut", "Échéance", "Montant dû", "Total versé", "Solde", "Date relance"])

    for r in relances:
        ws.append([
            f"{r.inscription.eleve.nom} {r.inscription.eleve.prenoms}",
            r.inscription.statut,
            f"{ r.echeance.nom } - { r.echeance.modalite.nom }",
            r.echeance_montant,
            r.total_verse,
            r.total_solde,
            r.date_relance.strftime('%d/%m/%Y')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relances_non_ajournees.xlsx'
    wb.save(response)
    return response

@fonctionnalite_autorisee('liste_relances_classe_non_a_jour')
def liste_relances_classe_non_a_jour(request, classe_id):
    classe = get_object_or_404(Classes, id=classe_id)
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    etablissement = request.user.etablissement
    aujourd_hui = date.today()

    # Filtrer les relances non soldées dont la date de relance est passée ou aujourd'hui
    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        inscription__classe=classe.id,
        statut='active',
        date_relance__lte=aujourd_hui
    ).select_related('inscription__eleve', 'echeance')

    return render(request, 'frontoffice/relances/liste_relances_classe_non_a_jour.html', {
        'relances': relances,
        'aujourd_hui': aujourd_hui,
        'classe': classe,  # 👈 à ajouter
    })


def export_relances_classe_pdf(request, classe_id):
    classe = get_object_or_404(Classes, id=classe_id)
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    etablissement = request.user.etablissement
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        inscription__classe=classe.id,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance')

    template = get_template('frontoffice/relances/export_relances_classe_pdf.html')
    html_string = template.render({
        'relances': relances,
        'date_du_jour': today,
        'etablissement': etablissement,
        'annee': annee_active,
        'classe': classe,  # 👈 à ajouter
    })

    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relances_non_ajournees.pdf"'
    return response


def export_relances_classe_excel(request, classe_id):
    classe = get_object_or_404(Classes, id=classe_id)
    annee_active = get_object_or_404(AnneeScolaires, active=True) 
    etablissement = request.user.etablissement
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        inscription__classe__etablissement=etablissement.id,
        inscription__classe=classe.id,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relances non à jour"

    # En-têtes
    ws.append(["Élève", "Statut", "Échéance", "Montant dû", "Total versé", "Solde", "Date relance"])

    for r in relances:
        ws.append([
            f"{r.inscription.eleve.nom} {r.inscription.eleve.prenoms}",
            r.inscription.statut,
            f"{ r.echeance.nom } - { r.echeance.modalite.nom }",
            r.echeance_montant,
            r.total_verse,
            r.total_solde,
            r.date_relance.strftime('%d/%m/%Y')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relances_non_ajournees.xlsx'
    wb.save(response)
    return response
    
#  les relances ---------------------------------------------------------------------------------------------------------------------

#  RAPPORT BACK OFFICE ---------------------------------------------------------------------------------------------------------------------
from django.db.models import Sum

def liste_relances_secretariat_non_a_jour(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    aujourd_hui = date.today()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        statut='active',
        date_relance__lte=aujourd_hui
    ).select_related('inscription__eleve', 'echeance', 'inscription__classe', 'inscription__classe__etablissement')

    return render(request, 'backoffice/relances/liste_relances_non_a_jour.html', {
        'relances': relances,
        'aujourd_hui': aujourd_hui,
    })

from django.contrib import messages
from django.views.decorators.http import require_POST

def envoyer_sms(numero, message):
    print(f"SMS envoyé à {numero} : {message}")  # remplacer par appel API réel

@require_POST
def envoyer_sms_relances(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    aujourd_hui = date.today()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        statut='active',
        date_relance__lte=aujourd_hui
    ).select_related('inscription__eleve__parent')

    total_sms = 0
    for relance in relances:
        eleve = relance.inscription.eleve
        parent = eleve.parent
        if parent and parent.telephone:
            message = (
                f"Bonjour, votre enfant {eleve.nom} {eleve.prenoms} a un retard de paiement. "
                f"Merci de régulariser la situation. Échéance : {relance.echeance.nom}. Montant : {relance.total_solde}F."
            )
            # Appel vers ton service d’envoi de SMS ici
            envoyer_sms(parent.telephone, message)
            total_sms += 1

    if not request.user.role.nom in ['Secrétaire Exécutif', 'Trésorerie']:
        messages.error(request, "Vous n'avez pas l'autorisation d’envoyer des SMS.")
    return redirect('liste_relances_secretariat_non_a_jour')
    #messages.success(request, f"{total_sms} SMS envoyés avec succès aux parents.")
    #return redirect('liste_relances_secretariat_non_a_jour')


def export_relances_secretariat_pdf(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True)
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance', 'inscription__classe__etablissement')

    template = get_template('backoffice/relances/export_relances_pdf.html')
    html_string = template.render({
        'relances': relances,
        'date_du_jour': today,
        'annee': annee_active,
    })

    pdf_file = HTML(string=html_string).write_pdf()
    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relances_non_ajournees.pdf"'
    return response

def export_relances_secretariat_excel(request):
    annee_active = get_object_or_404(AnneeScolaires, active=True) 
    today = now().date()

    relances = Relances.objects.filter(
        inscription__annee_scolaire=annee_active,
        statut='active',
        date_relance__lte=today
    ).select_related('inscription__eleve', 'echeance', 'inscription__classe__etablissement')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relances non à jour"

    ws.append(["Établissement", "Élève", "Statut", "Échéance", "Montant dû", "Total versé", "Solde", "Date relance"])

    for r in relances:
        ws.append([
            r.inscription.classe.etablissement.nom,
            f"{r.inscription.eleve.nom} {r.inscription.eleve.prenoms}",
            r.inscription.statut,
            f"{r.echeance.nom} - {r.echeance.modalite.nom}",
            r.echeance_montant,
            r.total_verse,
            r.total_solde,
            r.date_relance.strftime('%d/%m/%Y')
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relances_non_ajournees.xlsx'
    wb.save(response)
    return response



def tableau_relances_par_etablissement(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    resultats = (
        Relances.objects
        .filter(inscription__annee_scolaire=annee_active)
        .values('inscription__classe__etablissement__nom')
        .annotate(
            total_echeance=Sum('echeance_montant'),
            total_verse=Sum('total_verse'),
            total_solde=Sum('total_solde')
        )
        .order_by('inscription__classe__etablissement__nom')
    )

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }

    return render(request, 'backoffice/rapports/tableau_relances.html', context)


import openpyxl
from django.http import HttpResponse

def rapport_export_relances_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    relances = (
        Relances.objects
        .filter(inscription__annee_scolaire=annee_active)
        .values('inscription__classe__etablissement__nom')
        .annotate(
            somme_montant=Sum('echeance_montant'),
            somme_verse=Sum('total_verse'),
            somme_solde=Sum('total_solde'),
        )
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relances"

    headers = ["Établissement", "Montant Échéances", "Total Versé", "Total Solde"]
    sheet.append(headers)

    for item in relances:
        sheet.append([
            item['inscription__classe__etablissement__nom'],
            item['somme_montant'],
            item['somme_verse'],
            item['somme_solde'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relances_par_etablissement.xlsx"'
    workbook.save(response)
    return response



from django.template.loader import render_to_string
from django.http import HttpResponse
import weasyprint

def rapport_export_relances_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    relances = (
        Relances.objects
        .filter(inscription__annee_scolaire=annee_active)
        .values('inscription__classe__etablissement__nom')
        .annotate(
            somme_montant=Sum('echeance_montant'),
            somme_verse=Sum('total_verse'),
            somme_solde=Sum('total_solde'),
        )
    )

    html_string = render_to_string("backoffice/rapports/export_pdf.html", {
        'relances': relances,
        'annee': annee_active.libelle,
    })

    pdf_file = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="relances_par_etablissement.pdf"'
    return response

#  les etat des scolarites par classes pour un etablissement ------------------------------------------------------------------------------------------------
@fonctionnalite_autorisee('tableau_etatscolarite_pour_etablissement_classe')   
def tableau_etatscolarite_pour_etablissement_classe(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active)

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'frontoffice/rapports/tableau_relances_etablissement.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_par_etablissement_secretariat')
def tableau_etatscolarite_par_etablissement_secretariat(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissements = Etablissements.objects.all()

    resultats = []

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(classe__etablissement=etab, annee_scolaire=annee_active)

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'backoffice/rapports/tableau_relances_etablissement_secretariat.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_pour_etablissement_classe_abandons') 
def tableau_etatscolarite_pour_etablissement_classe_abandons(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, etat="abandon")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'frontoffice/rapports/tableau_relances_etablissement.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_pour_secretariat_classe_abandons')
def tableau_etatscolarite_pour_secretariat_classe_abandons(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissements = Etablissements.objects.all()

    resultats = []

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(classe__etablissement=etab, annee_scolaire=annee_active, etat="abandon")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'backoffice/rapports/tableau_relances_etablissement_abandons.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_pour_etablissement_classe_presents') 
def tableau_etatscolarite_pour_etablissement_classe_presents(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, etat="present")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'frontoffice/rapports/tableau_relances_etablissement.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_pour_secretariat_classe_presents')
def tableau_etatscolarite_pour_secretariat_classe_presents(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissements = Etablissements.objects.all()

    resultats = []

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(classe__etablissement=etab, annee_scolaire=annee_active, etat="present")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'backoffice/rapports/tableau_relances_etablissement.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_pour_etablissement_classe_affecte') 
def tableau_etatscolarite_pour_etablissement_classe_affecte(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="affecte")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'frontoffice/rapports/tableau_relances_affecte_etablissement.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_affecte_par_secretariat')
def tableau_etatscolarite_affecte_par_secretariat(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    
    # Cycles cibles
    cycles_cibles = ['Collège', 'Lycée', 'Technique']
    
    # Établissements qui ont au moins un de ces cycles
    etablissements = Etablissements.objects.filter(types__nom__in=cycles_cibles).distinct()
    
    resultats = []

    for etablissement in etablissements:
        classes = Classes.objects.filter(etablissement=etablissement)
        
        total_echeance = 0
        total_verse = 0

        for classe in classes:
            inscriptions = Inscriptions.objects.filter(
                classe=classe,
                annee_scolaire=annee_active,
                statut="affecte"
            )
            
            total_echeance += sum(ins.montant_total_du() for ins in inscriptions)
            total_verse += sum(ins.montant_total_paye() for ins in inscriptions)

        solde = total_echeance - total_verse

        resultats.append({
            'etablissement': etablissement,
            'total_echeance': total_echeance,
            'total_verse': total_verse,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }

    return render(request, 'backoffice/rapports/tableau_relances_affecte_tous_etablissements.html', context)

@fonctionnalite_autorisee('tableau_etatscolarite_nonaffecte_par_secretariat')
def tableau_etatscolarite_nonaffecte_par_secretariat(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    
    # Cycles cibles
    cycles_cibles = ['Collège', 'Lycée', 'Technique']
    
    # Établissements qui ont au moins un de ces cycles
    etablissements = Etablissements.objects.filter(types__nom__in=cycles_cibles).distinct()
    
    resultats = []

    for etablissement in etablissements:
        classes = Classes.objects.filter(etablissement=etablissement)
        
        total_echeance = 0
        total_verse = 0

        for classe in classes:
            inscriptions = Inscriptions.objects.filter(
                classe=classe,
                annee_scolaire=annee_active,
                statut="non_affecte"
            )
            
            total_echeance += sum(ins.montant_total_du() for ins in inscriptions)
            total_verse += sum(ins.montant_total_paye() for ins in inscriptions)

        solde = total_echeance - total_verse

        resultats.append({
            'etablissement': etablissement,
            'total_echeance': total_echeance,
            'total_verse': total_verse,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }

    return render(request, 'backoffice/rapports/tableau_relances_nonaffecte_tous_etablissements.html', context)

import tempfile
def rapport_export_etatscolarite_secretariat_affecte_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    resultats = []
    cycle_filtre = ['Collège', 'Lycée', 'Technique']
    etablissements = Etablissements.objects.filter(types__nom__in=cycle_filtre).distinct()

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(
            classe__etablissement=etab,
            annee_scolaire=annee_active,
            statut='affecte'
        ).select_related('classe')

        scolarite_attendue = sum(ins.montant_total_du() for ins in inscriptions)
        scolarite_payee = sum(ins.montant_total_paye() for ins in inscriptions)
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    html_string = render_to_string("backoffice/rapports/etatscolarite_affecte_secretariat_pdf.html", {
        'resultats': resultats,
        'annee_active': annee_active
    })

    with tempfile.NamedTemporaryFile(delete=True) as output:
        HTML(string=html_string).write_pdf(target=output.name)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="etat_scolarite_affectes.pdf"'
        return response

from openpyxl.styles import Font
def rapport_export_etatscolarite_secretariat_affecte_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    cycle_filtre = ['Collège', 'Lycée', 'Technique']
    etablissements = Etablissements.objects.filter(types__nom__in=cycle_filtre).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État Scolarité Élèves Affectés"

    # En-têtes
    headers = ["Établissement", "Total Scolarité", "Total Versé", "Total Solde"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Données
    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(
            classe__etablissement=etab,
            annee_scolaire=annee_active,
            statut="affecte"
        )

        total_du = sum(ins.montant_total_du() for ins in inscriptions)
        total_paye = sum(ins.montant_total_paye() for ins in inscriptions)
        solde = total_du - total_paye

        ws.append([
            etab.nom,
            float(total_du),
            float(total_paye),
            float(solde)
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"etat_scolarite_affectes_{annee_active}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

import tempfile
def rapport_export_etatscolarite_secretariat_nonaffecte_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    resultats = []
    cycle_filtre = ['Collège', 'Lycée', 'Technique']
    etablissements = Etablissements.objects.filter(types__nom__in=cycle_filtre).distinct()

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(
            classe__etablissement=etab,
            annee_scolaire=annee_active,
            statut='non_affecte'
        ).select_related('classe')

        scolarite_attendue = sum(ins.montant_total_du() for ins in inscriptions)
        scolarite_payee = sum(ins.montant_total_paye() for ins in inscriptions)
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    html_string = render_to_string("backoffice/rapports/etatscolarite_nonaffecte_secretariat_pdf.html", {
        'resultats': resultats,
        'annee_active': annee_active
    })

    with tempfile.NamedTemporaryFile(delete=True) as output:
        HTML(string=html_string).write_pdf(target=output.name)
        output.seek(0)
        response = HttpResponse(output.read(), content_type='application/pdf')
        response['Content-Disposition'] = 'inline; filename="etat_scolarite_non_affectes.pdf"'
        return response

from openpyxl.styles import Font
def rapport_export_etatscolarite_secretariat_nonaffecte_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    cycle_filtre = ['Collège', 'Lycée', 'Technique']
    etablissements = Etablissements.objects.filter(types__nom__in=cycle_filtre).distinct()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "État Scolarité Élèves Affectés"

    # En-têtes
    headers = ["Établissement", "Total Scolarité", "Total Versé", "Total Solde"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Données
    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(
            classe__etablissement=etab,
            annee_scolaire=annee_active,
            statut="non_affecte"
        )

        total_du = sum(ins.montant_total_du() for ins in inscriptions)
        total_paye = sum(ins.montant_total_paye() for ins in inscriptions)
        solde = total_du - total_paye

        ws.append([
            etab.nom,
            float(total_du),
            float(total_paye),
            float(solde)
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"etat_scolarite_non_affectes_{annee_active}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

@fonctionnalite_autorisee('tableau_etatscolarite_pour_etablissement_classe_nonaffecte')  
def tableau_etatscolarite_pour_etablissement_classe_nonaffecte(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="non_affecte")

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    context = {
        'annee_active': annee_active,
        'resultats': resultats
    }
    return render(request, 'frontoffice/rapports/tableau_relances_nonaffecte_etablissement.html', context)

import openpyxl
from django.http import HttpResponse

def rapport_export_etatscolarite_classe_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active)
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "État scolarité par classe"

    headers = ["Classe", "Montant Échéances", "Total Versé", "Total Solde"]
    sheet.append(headers)

    for item in resultats:
        sheet.append([
            item['classe'],
            item['total_echeance'],
            item['total_verse'],
            item['total_solde'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.xlsx"'
    workbook.save(response)
    return response

def rapport_export_etatscolarite_classe_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active)
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    html_string = render_to_string("frontoffice/rapports/etatscolarite_export_pdf.html", {
       'annee_active': annee_active,
       'resultats': resultats
    })

    pdf_file = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.pdf"'
    return response

def rapport_export_etatscolarite_secretariat_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissements = Etablissements.objects.all()

    resultats = []

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(classe__etablissement=etab, annee_scolaire=annee_active)

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })


    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "État scolarité par etablissement"

    headers = ["Établissement", "Montant Échéances", "Total Versé", "Total Solde"]
    sheet.append(headers)

    for item in resultats:
        sheet.append([
            item['etablissement'],
            item['total_echeance'],
            item['total_verse'],
            item['total_solde'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.xlsx"'
    workbook.save(response)
    return response

def rapport_export_etatscolarite_secretariat_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissements = Etablissements.objects.all()

    resultats = []

    for etab in etablissements:
        inscriptions = Inscriptions.objects.filter(classe__etablissement=etab, annee_scolaire=annee_active)

        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'etablissement': etab,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })


    html_string = render_to_string("backoffice/rapports/etatscolarite_secretariat_export_pdf.html", {
       'annee_active': annee_active,
       'resultats': resultats
    })

    pdf_file = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.pdf"'
    return response


def rapport_export_etatscolarite_classe_affecte_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="affecte")
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "État scolarité par classe"

    headers = ["Classe", "Montant Échéances", "Total Versé", "Total Solde"]
    sheet.append(headers)

    for item in resultats:
        sheet.append([
            item['classe'],
            item['total_echeance'],
            item['total_verse'],
            item['total_solde'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.xlsx"'
    workbook.save(response)
    return response

def rapport_export_etatscolarite_classe_affecte_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="affecte")
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    html_string = render_to_string("frontoffice/rapports/etatscolarite_export_pdf.html", {
       'annee_active': annee_active,
       'resultats': resultats
    })

    pdf_file = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe.pdf"'
    return response

def rapport_export_etatscolarite_classe_non_affecte_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="non_affecte")
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "État scolarité par classe"

    headers = ["Classe", "Montant Échéances", "Total Versé", "Total Solde"]
    sheet.append(headers)

    for item in resultats:
        sheet.append([
            item['classe'],
            item['total_echeance'],
            item['total_verse'],
            item['total_solde'],
        ])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_non_affecte_classe.xlsx"'
    workbook.save(response)
    return response

def rapport_export_etatscolarite_classe_non_affecte_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    classes = Classes.objects.filter(etablissement=etablissement)
    resultats = []

    for c in classes:
        inscriptions = Inscriptions.objects.filter(classe=c, annee_scolaire=annee_active, statut="non_affecte")
        scolarite_attendue = sum([ins.montant_total_du() for ins in inscriptions])
        scolarite_payee = sum([ins.montant_total_paye() for ins in inscriptions])
        solde = scolarite_attendue - scolarite_payee

        resultats.append({
            'classe': c.nom,
            'total_echeance': scolarite_attendue,
            'total_verse': scolarite_payee,
            'total_solde': solde
        })

    html_string = render_to_string("frontoffice/rapports/etatscolarite_non_affecte_export_pdf.html", {
       'annee_active': annee_active,
       'resultats': resultats
    })

    pdf_file = weasyprint.HTML(string=html_string).write_pdf()

    response = HttpResponse(pdf_file, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="etat_scolarite_par_classe_non_affecte.pdf"'
    return response

# effectif -----------------------------------------------------------------------------------------------------------
from django.shortcuts import render
from django.db.models import Count
from .models import Inscriptions, AnneeScolaires

@fonctionnalite_autorisee('effectif_par_niveau_genre')
def effectif_par_niveau_genre(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    if not annee_active:
        return render(request, 'backoffice/rapports/effectif_global_etab.html', {
            'message': "Aucune année scolaire active."
        })
 
    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Préscolaire", "Primaire"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(effectif=Count('id'))
    )

    niveaux = sorted({item['classe__niveau__nom'] for item in data})
    effectif_dict = {}
    total_par_niveau_genre = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}
    total_garcons = 0
    total_filles = 0

    for etab in etablissements_filtres:
        effectif_dict[etab.nom] = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}

    for item in data:
        etab = item['classe__etablissement__nom']
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif = item['effectif']

        effectif_dict[etab][niveau][sexe] = effectif
        total_par_niveau_genre[niveau][sexe] += effectif
        if sexe == 'M':
            total_garcons += effectif
        elif sexe == 'F':
            total_filles += effectif

    for etab, niveaux_data in effectif_dict.items():
        total_g = sum(v['M'] for v in niveaux_data.values())
        total_f = sum(v['F'] for v in niveaux_data.values())
        effectif_dict[etab]['TOTAL'] = {
            'M': total_g,
            'F': total_f,
            'TOTAL': total_g + total_f
        }

    context = {
        'annee': annee_active,
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'total_par_niveau_genre': total_par_niveau_genre,
        'total_garcons': total_garcons,
        'total_filles': total_filles,
    }

    return render(request, 'backoffice/rapports/effectif_global_etab.html', context)

@fonctionnalite_autorisee('effectif_par_niveau_genre_sec')
def effectif_par_niveau_genre_sec(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    if not annee_active:
        return render(request, 'backoffice/rapports/effectif_global_etab_sec.html', {
            'message': "Aucune année scolaire active."
        })

    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Collège", "Lycée", "Technique"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(effectif=Count('id'))
    )

    niveaux = sorted({item['classe__niveau__nom'] for item in data})
    effectif_dict = {}
    total_par_niveau_genre = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}
    total_garcons = 0
    total_filles = 0

    for etab in etablissements_filtres:
        effectif_dict[etab.nom] = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}

    for item in data:
        etab = item['classe__etablissement__nom']
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif = item['effectif']

        effectif_dict[etab][niveau][sexe] = effectif
        total_par_niveau_genre[niveau][sexe] += effectif
        if sexe == 'M':
            total_garcons += effectif
        elif sexe == 'F':
            total_filles += effectif

    for etab, niveaux_data in effectif_dict.items():
        total_g = sum(v['M'] for v in niveaux_data.values())
        total_f = sum(v['F'] for v in niveaux_data.values())
        effectif_dict[etab]['TOTAL'] = {
            'M': total_g,
            'F': total_f,
            'TOTAL': total_g + total_f
        }

    context = {
        'annee': annee_active,
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'total_par_niveau_genre': total_par_niveau_genre,
        'total_garcons': total_garcons,
        'total_filles': total_filles,
    }

    return render(request, 'backoffice/rapports/effectif_global_etab_sec.html', context)

# -----------------------------
from collections import defaultdict
from django.db.models import Count

def build_effectif_dict(data):
    effectif_dict = {}

    for item in data:
        etab = item['classe__etablissement__nom']
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        total = item['total']

        effectif_dict.setdefault(etab, {}).setdefault(niveau, {'M': 0, 'F': 0})
        effectif_dict[etab][niveau][sexe] = total

    # Calcul des totaux par établissement
    for etab, niveaux_data in effectif_dict.items():
        total_m = sum(niv.get('M', 0) for niv in niveaux_data.values())
        total_f = sum(niv.get('F', 0) for niv in niveaux_data.values())
        niveaux_data['TOTAL'] = {
            'M': total_m,
            'F': total_f,
            'TOTAL': total_m + total_f
        }

    return effectif_dict


import openpyxl
from django.http import HttpResponse

def export_effectif_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Préscolaire", "Primaire"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    effectif_dict = build_effectif_dict(data)

    # Extraire dynamiquement les niveaux présents dans les données triés par nom
    niveaux_noms = sorted({item['classe__niveau__nom'] for item in data})
    niveaux = list(Niveaux.objects.filter(nom__in=niveaux_noms).order_by('nom'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Effectifs par niveau et genre"

    # En-tête
    header = ["Établissement"]
    for niveau in niveaux:
        header.append(f"{niveau.nom} G")
        header.append(f"{niveau.nom} F")
    header.extend(["Total G", "Total F", "Total"])
    ws.append(header)

    for etab, niveaux_data in effectif_dict.items():
        row = [etab]
        for niveau in niveaux:
            data_niveau = niveaux_data.get(niveau.nom, {})
            row.append(data_niveau.get('M', 0))
            row.append(data_niveau.get('F', 0))
        row.append(niveaux_data['TOTAL']['M'])
        row.append(niveaux_data['TOTAL']['F'])
        row.append(niveaux_data['TOTAL']['TOTAL'])
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=effectif_par_niveau_genre.xlsx'
    wb.save(response)
    return response

def export_effectif_excel_sec(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Collège", "Lycée", "Technique"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    effectif_dict = build_effectif_dict(data)

    # Extraire dynamiquement les niveaux présents dans les données triés par nom
    niveaux_noms = sorted({item['classe__niveau__nom'] for item in data})
    niveaux = list(Niveaux.objects.filter(nom__in=niveaux_noms).order_by('nom'))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Effectifs par niveau et genre"

    # En-tête
    header = ["Établissement"]
    for niveau in niveaux:
        header.append(f"{niveau.nom} G")
        header.append(f"{niveau.nom} F")
    header.extend(["Total G", "Total F", "Total"])
    ws.append(header)

    for etab, niveaux_data in effectif_dict.items():
        row = [etab]
        for niveau in niveaux:
            data_niveau = niveaux_data.get(niveau.nom, {})
            row.append(data_niveau.get('M', 0))
            row.append(data_niveau.get('F', 0))
        row.append(niveaux_data['TOTAL']['M'])
        row.append(niveaux_data['TOTAL']['F'])
        row.append(niveaux_data['TOTAL']['TOTAL'])
        ws.append(row)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=effectif_par_niveau_genre.xlsx'
    wb.save(response)
    return response

    

# views.py
from django.template.loader import get_template
from xhtml2pdf import pisa
from io import BytesIO
from django.http import HttpResponse

def build_total_par_niveau_genre(data):
    total_par_niveau = defaultdict(lambda: {'M': 0, 'F': 0})

    total_general_m = 0
    total_general_f = 0

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        total = item['total']
        total_par_niveau[niveau][sexe] += total

        if sexe == 'M':
            total_general_m += total
        elif sexe == 'F':
            total_general_f += total

    total_par_niveau['TOTAL'] = {
        'M': total_general_m,
        'F': total_general_f,
        'TOTAL': total_general_m + total_general_f
    }

    return total_par_niveau



def export_effectif_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Préscolaire", "Primaire"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )
    total_par_niveau_genre = build_total_par_niveau_genre(data)

    effectif_dict = build_effectif_dict(data)

    # Extraire dynamiquement les niveaux présents dans les données triés par nom
    niveaux_noms = sorted({item['classe__niveau__nom'] for item in data})
    niveaux = list(Niveaux.objects.filter(nom__in=niveaux_noms).order_by('nom'))

    template = get_template('backoffice/rapports/pdf_template_effectif.html')
    html = template.render({
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'annee': annee_active,
        'total_par_niveau_genre': total_par_niveau_genre,  # <-- Ajout ici
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=effectif_par_niveau_genre.pdf'
    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur de génération du PDF', status=500)
    return response

def export_effectif_pdf_sec(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    etablissements_filtres = Etablissements.objects.filter(
        types__nom__in=["Collège", "Lycée", "Technique"]
    ).distinct()

    data = (
        Inscriptions.objects
        .filter(
            annee_scolaire=annee_active,
            classe__etablissement__in=etablissements_filtres
        )
        .values('classe__etablissement__nom', 'classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )
    total_par_niveau_genre = build_total_par_niveau_genre(data)

    effectif_dict = build_effectif_dict(data)

    # Extraire dynamiquement les niveaux présents dans les données triés par nom
    niveaux_noms = sorted({item['classe__niveau__nom'] for item in data})
    niveaux = list(Niveaux.objects.filter(nom__in=niveaux_noms).order_by('nom'))

    template = get_template('backoffice/rapports/pdf_template_effectif.html')
    html = template.render({
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'annee': annee_active,
        'total_par_niveau_genre': total_par_niveau_genre,  # <-- Ajout ici
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=effectif_par_niveau_genre.pdf'
    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur de génération du PDF', status=500)
    return response

# ====== Voir l'effectif pour un etablissement =============================================================================
from django.shortcuts import render
from django.db.models import Count

@fonctionnalite_autorisee('effectif_par_niveau_genre_etablissement')  
def effectif_par_niveau_genre_etablissement(request):
    # Année active
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    if not annee_active:
        return render(request, 'frontoffice/rapports/effectif_global_etablissement.html', {
            'message': 'Aucune année scolaire active.',
        })

    # Établissement de l'utilisateur connecté
    etablissement = request.user.etablissement

    # Regroupement par niveau et sexe
    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement)
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(effectif=Count('id'))
    )

    # Préparation des niveaux uniques
    niveaux = sorted({item['classe__niveau__nom'] for item in data})
    effectif_dict = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}

    # Remplissage des effectifs
    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['effectif']

    # Calcul total
    total_garcons = sum(v['M'] for v in effectif_dict.values())
    total_filles = sum(v['F'] for v in effectif_dict.values())

    context = {
        'annee': annee_active,
        'etablissement': etablissement,
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'total_garcons': total_garcons,
        'total_filles': total_filles,
        'total': total_garcons + total_filles,
    }

    return render(request, 'frontoffice/rapports/effectif_global_etablissement.html', context)

def effectif_par_niveau_genre_etablissementU(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement)
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(effectif=Count('id'))
    )
    
    niveaux = sorted({item['classe__niveau__nom'] for item in data})
    effectif_dict = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['effectif']

    context = {
        'annee': annee_active,
        'etablissement': etablissement,
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
    }

    return render(request, 'frontoffice/rapports/effectif_global_etablissement.html', context)

@fonctionnalite_autorisee('effectif_par_niveau_genre_abandon_etablissement') 
def effectif_par_niveau_genre_abandon_etablissement(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement, etat="abandon")
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(effectif=Count('id'))
    )
    
    # Préparation des niveaux uniques
    niveaux = sorted({item['classe__niveau__nom'] for item in data})
    effectif_dict = {niveau: {'M': 0, 'F': 0} for niveau in niveaux}

    # Remplissage des effectifs
    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['effectif']

    # Calcul total
    total_garcons = sum(v['M'] for v in effectif_dict.values())
    total_filles = sum(v['F'] for v in effectif_dict.values())

    context = {
        'annee': annee_active,
        'etablissement': etablissement,
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'total_garcons': total_garcons,
        'total_filles': total_filles,
        'total': total_garcons + total_filles,
    }

    return render(request, 'frontoffice/rapports/effectif_global_etablissement_abandon.html', context)

# back  abandons  -------------------------------------------------------------------------------------------------------------------
from django.db.models import Count, F
@fonctionnalite_autorisee('effectif_par_niveau_genre_abandon_tous_etablissements')
def effectif_par_niveau_genre_abandon_tous_etablissements(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, etat="abandon")
        .values(etablissement=F('classe__etablissement__nom'), niveau=F('classe__niveau__nom'))
        .annotate(effectif=Count('id'))
        .order_by('etablissement', 'niveau')
    )

    total_abandons = sum(item['effectif'] for item in data)

    context = {
        'annee': annee_active,
        'data': data,  # liste de dictionnaires : [{'etablissement': ..., 'niveau': ..., 'effectif': ...}, ...]
        'total_abandons': total_abandons,
    }

    return render(request, 'backoffice/rapports/effectif_abandon_par_etablissement.html', context)

def export_effectif_abandon_global_excel(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, etat="abandon")
        .values(etablissement=F('classe__etablissement__nom'), niveau=F('classe__niveau__nom'))
        .annotate(effectif=Count('id'))
        .order_by('etablissement', 'niveau')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Abandons Global"

    ws.append(["Établissement", "Niveau", "Effectif d'abandon"])

    for item in data:
        ws.append([item['etablissement'], item['niveau'], item['effectif']])

    ws.append(["", "Total général", sum(item['effectif'] for item in data)])

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=effectif_abandon_global.xlsx'
    wb.save(response)
    return response


def export_effectif_abandon_global_pdf(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, etat="abandon")
        .values(etablissement=F('classe__etablissement__nom'), niveau=F('classe__niveau__nom'))
        .annotate(effectif=Count('id'))
        .order_by('etablissement', 'niveau')
    )

    total_abandons = sum(item['effectif'] for item in data)

    template = get_template("backoffice/rapports/pdf_effectif_abandon_global.html")
    html = template.render({
        "data": data,
        "annee": annee_active,
        "total_abandons": total_abandons,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=effectif_abandon_global.pdf'
    pisa.CreatePDF(html, dest=response)
    return response


def export_effectif_excel_etablissement(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement
    #niveaux = list(Niveaux.objects.all().order_by('nom'))
    # Cycles liés à l’établissement → donc niveaux de cet établissement
    cycles = etablissement.types.all()
    niveaux = list(Niveaux.objects.filter(cycle__in=cycles).order_by('nom'))

    # Initialisation du dictionnaire d'effectif
    effectif_dict = {niveau.nom: {'M': 0, 'F': 0} for niveau in niveaux}

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement)
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['total']

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Effectifs - {etablissement.nom}"

    # En-tête
    ws.append(["Niveau", "Garçons", "Filles", "Total"])

    for niveau in niveaux:
        nom = niveau.nom
        garçons = effectif_dict[nom]['M']
        filles = effectif_dict[nom]['F']
        total = garçons + filles
        ws.append([nom, garçons, filles, total])

    # Export
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=effectifs_{etablissement.nom}.xlsx'
    wb.save(response)
    return response


def export_effectif_pdf_etablissement(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement
    #niveaux = list(Niveaux.objects.all().order_by('nom'))
    # Cycles liés à l’établissement → donc niveaux de cet établissement
    cycles = etablissement.types.all()
    niveaux = list(Niveaux.objects.filter(cycle__in=cycles).order_by('nom'))

    # Initialisation du dictionnaire d'effectif
    effectif_dict = {niveau.nom: {'M': 0, 'F': 0} for niveau in niveaux}

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement)
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['total']

    template = get_template('frontoffice/rapports/pdf_template_effectif_etablissement.html')
    html = template.render({
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'annee': annee_active,
        'etablissement': etablissement,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=effectifs_{etablissement.nom}.pdf'

    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur de génération du PDF', status=500)
    return response

def export_effectif_excel_etablissement_abandon(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement
    #niveaux = list(Niveaux.objects.all().order_by('nom'))
    # Cycles liés à l’établissement → donc niveaux de cet établissement
    cycles = etablissement.types.all()
    niveaux = list(Niveaux.objects.filter(cycle__in=cycles).order_by('nom'))

    # Initialisation du dictionnaire d'effectif
    effectif_dict = {niveau.nom: {'M': 0, 'F': 0} for niveau in niveaux}

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement, etat="abandon")
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['total']

    # Création du fichier Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Effectifs - {etablissement.nom}"

    # En-tête
    ws.append(["Niveau", "Garçons", "Filles", "Total"])

    for niveau in niveaux:
        nom = niveau.nom
        garçons = effectif_dict[nom]['M']
        filles = effectif_dict[nom]['F']
        total = garçons + filles
        ws.append([nom, garçons, filles, total])

    # Export
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=effectifs_{etablissement.nom}.xlsx'
    wb.save(response)
    return response


def export_effectif_pdf_etablissement_abandon(request):
    annee_active = AnneeScolaires.objects.filter(active=True).first()
    etablissement = request.user.etablissement
    #niveaux = list(Niveaux.objects.all().order_by('nom'))
    # Cycles liés à l’établissement → donc niveaux de cet établissement
    cycles = etablissement.types.all()
    niveaux = list(Niveaux.objects.filter(cycle__in=cycles).order_by('nom'))

    # Initialisation du dictionnaire d'effectif
    effectif_dict = {niveau.nom: {'M': 0, 'F': 0} for niveau in niveaux}

    data = (
        Inscriptions.objects
        .filter(annee_scolaire=annee_active, classe__etablissement=etablissement, etat="abandon")
        .values('classe__niveau__nom', 'eleve__sexe')
        .annotate(total=Count('id'))
    )

    for item in data:
        niveau = item['classe__niveau__nom']
        sexe = item['eleve__sexe']
        effectif_dict[niveau][sexe] = item['total']

    template = get_template('frontoffice/rapports/pdf_template_effectif_etablissement_abandon.html')
    html = template.render({
        'niveaux': niveaux,
        'effectif_dict': effectif_dict,
        'annee': annee_active,
        'etablissement': etablissement,
    })

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename=effectifs_{etablissement.nom}.pdf'

    pisa_status = pisa.CreatePDF(BytesIO(html.encode('utf-8')), dest=response)
    if pisa_status.err:
        return HttpResponse('Erreur de génération du PDF', status=500)
    return response

# ============= ESPACE PARENT ==========================================================================================================================
from django.shortcuts import get_object_or_404

def get_parent(request):
    return get_object_or_404(Parents, utilisateur=request.user)

def liste_enfants(request):
    parent = get_parent(request)
    enfants = Eleves.objects.filter(parent=parent)
    return render(request, 'espace_parent/liste_enfants.html', {'enfants': enfants})

@fonctionnalite_autorisee('detail_eleve_parent')
def detail_eleve_parent(request, matricule):
    eleve = get_object_or_404(Eleves, matricule=matricule)
    inscriptions = Inscriptions.objects.filter(eleve=eleve).select_related('classe', 'annee_scolaire')
    lienparent = LienParente.objects.filter(eleve__matricule=matricule).select_related('parent__utilisateur')
    paiements = Paiements.objects.filter(inscription__eleve=eleve).select_related('inscription')
    paiementtransports = PaiementsTransports.objects.filter(inscription__eleve=eleve).select_related('inscription')
    paiementcantines = PaiementsCantines.objects.filter(inscription__eleve=eleve).select_related('inscription')
    notes = Notes.objects.filter(eleve=eleve).select_related('matiere', 'periode')
    inscription_active = inscriptions.first()  # ou filtre selon l'année active 
    print(lienparent)
    
    #classe = get_object_or_404(Classes, id=inscription_active.classe.id)
    #inscription_active = eleve.inscriptions.filter(active=True).first()

    if not inscription_active:
        # Redirige ou affiche une page d'erreur personnalisée
        messages.error(request, "Aucune inscription active trouvée pour cet élève.")
        return redirect('espace_parent')  # À adapter

    classe = get_object_or_404(Classes, id=inscription_active.classe.id)
    
    emplois = EmploiTemps.objects.filter(classe=classe).order_by('jour', 'heure_debut')

    jours = ['lundi', 'mardi', 'mercredi', 'jeudi', 'vendredi', 'samedi']
    
    # Construire toutes les heures de début possibles
    heures = sorted(set(emploi.heure_debut.strftime('%H:%M') for emploi in emplois))
    heuresf = sorted(set(emploi.heure_fin.strftime('%H:%M') for emploi in emplois))

    #print(emplois) 
    #print(jours)
    #print(heures)
    # Afficher les matières et les professeurs pour chaque emploi
    #for emploi in emplois:
    #    print(f"Matière: {emploi.matiere.nom}")
        #print(f"Professeur: {emploi.professeur.nom_complet}")
    
    context = {
        'eleve': eleve,
        'inscriptions': inscriptions,
        'parents': lienparent,
        'paiements': paiements,
        'paiements_transport' : paiementtransports,
        'paiements_cantine' : paiementcantines,
        'notes': notes,
        'inscription_active': inscription_active,
        
        'classe': classe,
        'emplois': emplois,
        'jours': jours,
        'heures': heures,
        'heuresf': heuresf,
    }
    return render(request, 'espace_parent/detail_eleve.html', context)

@fonctionnalite_autorisee('fiche_eleve')   
def fiche_eleve(request, eleve_id):
    parent = get_parent(request)
    eleve = get_object_or_404(Eleves, id=eleve_id, parent=parent)
    return render(request, 'parents/fiche_eleve.html', {'eleve': eleve})

@fonctionnalite_autorisee('paiements_eleve')   
def paiements_eleve(request):
    parent = get_parent(request)
    enfants = Eleves.objects.filter(parent=parent)

    data = []
    for enfant in enfants:
        inscription = enfant.inscription_active()
        paiements = Paiements.objects.filter(inscription=inscription) if inscription else []
        data.append({'enfant': enfant, 'paiements': paiements, 'inscription': inscription})

    return render(request, 'espace_parent/paiements.html', {'data': data})

@fonctionnalite_autorisee('bulletins_eleve')   
def bulletins_eleve(request):
    parent = get_parent(request)
    enfants = Eleves.objects.filter(parent=parent)

    data = []
    for enfant in enfants:
        bulletins = Bulletins.objects.filter(eleve=enfant)
        data.append({'enfant': enfant, 'bulletins': bulletins})

    return render(request, 'espace_parent/bulletins.html', {'data': data})



















# ===========================================================                TEST               ==============================================================
def ajouter_eleveN(request):
    if request.method == 'POST':
        eleve_form = EleveForm(request.POST, request.FILES)
        parent_form = ParentForm(request.POST)
        formset = LienParenteFormSet(request.POST)

        if eleve_form.is_valid() and parent_form.is_valid() and formset.is_valid():
            try:
                telephone = parent_form.cleaned_data.get('telephone')
                email = parent_form.cleaned_data.get('email')
                nom_complet = parent_form.cleaned_data['nom_complet'].strip()
                nom_parts = nom_complet.split()
                prenom = nom_parts[0]
                nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''

                parent = Parents.objects.filter(telephone=telephone).first()
                if not parent:
                    # Vérifie aussi l'email pour éviter doublons
                    email = parent_form.cleaned_data.get('email')
                    parent = Parents.objects.filter(email=email).first()

                if not parent:
                    parent = parent_form.save()



                # Création ou récupération de l'utilisateur parent
                username = telephone
                password = telephone  # mot de passe = téléphone

                user, created = Utilisateurs.objects.get_or_create(username=username, defaults={
                    'first_name': prenom,
                    'last_name': nom,
                    'email': email or '',
                    'password': make_password(password),
                })

                # Création de l'élève
                eleve = eleve_form.save(commit=False)
                eleve.parent = parent
                eleve.save()

                formset.instance = eleve
                formset.save()

                # Envoi Email
                if parent.email:
                    send_mail(
                        subject="Informations d’accès à la plateforme",
                        message=(
                            f"Bonjour {parent.nom_complet},\n\n"
                            f"Votre enfant {eleve.nom} {eleve.prenoms} (matricule : {eleve.matricule}) "
                            f"a été inscrit avec succès.\n\n"
                            f"Identifiants de connexion parent :\n"
                            f"Nom d'utilisateur : {username}\n"
                            f"Mot de passe : {password}\n\n"
                            "Merci de vous connecter à la plateforme."
                        ),
                        from_email=settings.DEFAULT_FROM_EMAIL,
                        recipient_list=[parent.email],
                        fail_silently=True,
                    )

                return redirect('liste_eleves')
            except Exception as e:
                logger.error(f"Erreur lors de la création de l'élève ou du parent : {str(e)}")

        else:
            logger.warning("Formulaires invalides :")
            logger.warning(eleve_form.errors)
            logger.warning(parent_form.errors)
            logger.warning(formset.errors)

    else:
        eleve_form = EleveForm()
        parent_form = ParentForm()
        formset = LienParenteFormSet()

    return render(request, 'frontoffice/eleves/formulaire.html', {
        'form': eleve_form,
        'form_parent': parent_form,
        'formset': formset,
        'titre': "Ajouter un élève"
    })


def ajouter_eleve_t(request):
    if request.method == 'POST':
        eleve_form = EleveForm(request.POST, request.FILES)
        parent_form = ParentForm(request.POST)
        formset = LienParenteFormSet(request.POST)

        if eleve_form.is_valid() and parent_form.is_valid() and formset.is_valid():
            telephone = parent_form.cleaned_data.get('telephone')
            email = parent_form.cleaned_data.get('email')
            nom_complet = parent_form.cleaned_data['nom_complet'].strip()
            nom_parts = nom_complet.split()
            prenom = nom_parts[0]
            nom = " ".join(nom_parts[1:]) if len(nom_parts) > 1 else ''
            role = Roles.objects.get_or_create(nom="Parents")[0]
            
            parent = Parents.objects.filter(telephone=telephone).first()
            if not parent:
                parent = parent_form.save()

            # Création ou récupération de l'utilisateur parent
            username = telephone
            password = telephone  # mot de passe = téléphone

            user, created = Utilisateurs.objects.get_or_create(username=username, defaults={
                'first_name': prenom,
                'last_name': nom,
                'email': email or '',
                'password': make_password(password),
                'role': role,
                'telephone': telephone,
            })

            # Création de l'élève
            eleve = eleve_form.save(commit=False)
            eleve.parent = parent
            eleve.save()

            formset.instance = eleve
            formset.save()

            # Envoi Email
            #if parent.email:
            #    send_mail(
            #        subject="Informations d’accès à la plateforme",
            #        message=(
            #            f"Bonjour {parent.nom_complet},\n\n"
            #            f"Votre enfant {eleve.nom} {eleve.prenoms} (matricule : {eleve.matricule}) "
            #            f"a été inscrit avec succès.\n\n"
            #            f"Identifiants de connexion parent :\n"
            #            f"Nom d'utilisateur : {username}\n"
            #            f"Mot de passe : {password}\n\n"
            #            "Merci de vous connecter à la plateforme."
            #        ),
            #        from_email=settings.DEFAULT_FROM_EMAIL,
            #        recipient_list=[parent.email],
            #        fail_silently=True,
            #    )

            # Envoi SMS — à implémenter via une API comme Orange, Twilio, etc.

            return redirect('liste_eleves')

    else:
        eleve_form = EleveForm()
        parent_form = ParentForm()
        formset = LienParenteFormSet()

    return render(request, 'frontoffice/eleves/formulaire.html', {
        'form': eleve_form,
        'form_parent': parent_form,
        'formset': formset,
        'titre': "Ajouter un élève"
    })


def ajouter_eleve_ok(request):
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES)
        formset = LienParenteFormSet(request.POST)
        if form.is_valid() and formset.is_valid():
            eleve = form.save()
            formset.instance = eleve
            formset.save()
            return redirect('liste_eleves')
    else:
        form = EleveForm()
        formset = LienParenteFormSet()
    return render(request, 'frontoffice/eleves/formulaire.html', {'form': form, 'formset': formset, 'titre': "Ajouter un élève"})


def modifier_eleveOUI(request, eleve_id):
    eleve = get_object_or_404(Eleves, pk=eleve_id)
    if request.method == 'POST':
        form = EleveForm(request.POST, request.FILES, instance=eleve)
        formset = LienParenteFormSet(request.POST, instance=eleve)
        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()
            return redirect('liste_eleves')
    else:
        form = EleveForm(instance=eleve)
        formset = LienParenteFormSet(instance=eleve)
    return render(request, 'frontoffice/eleves/formulaire.html', {'form': form, 'formset': formset, 'titre': "Modifier l'élève"})